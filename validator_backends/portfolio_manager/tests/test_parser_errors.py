"""Carrier mismatch, corruption, shape, and workbook-bound tests.

Portfolio Manager downloads have historically used several Excel encodings and
occasionally misleading suffixes. These tests distinguish intentionally
supported compatibility cases from corrupted input while proving that row and
column bounds do not depend on optional OOXML worksheet dimensions.
"""

from __future__ import annotations

import io
import struct
import zipfile
from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import Workbook

from validator_backends.portfolio_manager import parser
from validator_backends.portfolio_manager.parser import (
    PortfolioManagerParseError,
    parse_report_bytes,
)
from validator_backends.portfolio_manager.tests.test_parser import _xlsx_bytes


ASSETS = Path(__file__).with_name("assets")


def _ooxml_with_part(name: str, payload: bytes = b"unsafe") -> bytes:
    """Add one prohibited package part to an otherwise valid workbook."""
    source = io.BytesIO(_xlsx_bytes())
    output = io.BytesIO()
    with (
        zipfile.ZipFile(source) as existing,
        zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as updated,
    ):
        for member in existing.infolist():
            updated.writestr(member, existing.read(member))
        updated.writestr(name, payload)
    return output.getvalue()


def test_unknown_extension_is_rejected_before_content_sniffing() -> None:
    """Workflow file-type mistakes must not silently broaden accepted carriers."""
    with pytest.raises(PortfolioManagerParseError, match="Unsupported report extension"):
        parse_report_bytes(b"Portfolio Manager", filename="report.csv")


def test_corrupt_ooxml_is_reported_as_a_parse_error() -> None:
    """A ZIP signature alone cannot make truncated content a valid workbook."""
    with pytest.raises(PortfolioManagerParseError, match="Could not read OOXML"):
        parse_report_bytes(b"PK\x03\x04truncated", filename="report.xlsx")


def test_corrupt_biff_is_reported_as_a_parse_error() -> None:
    """A legacy suffix cannot route arbitrary bytes into row normalization."""
    with pytest.raises(PortfolioManagerParseError, match="Could not read legacy XLS"):
        parse_report_bytes(b"not an OLE workbook", filename="report.xls")


def test_ooxml_with_legacy_xls_suffix_is_intentionally_supported() -> None:
    """Some download paths retain an XLS name while returning an OOXML package."""
    records = parse_report_bytes(_xlsx_bytes(), filename="report.xls")

    assert records[0].property_id == "00123"
    assert records[0].carrier == "xls"


def test_biff_with_xlsx_suffix_is_rejected_as_mislabeled() -> None:
    """Modern XLSX submissions must contain OOXML rather than a legacy BIFF file."""
    content = (ASSETS / "portfolio-manager-custom-report-single-anonymized.xls").read_bytes()

    with pytest.raises(PortfolioManagerParseError, match="Could not read OOXML"):
        parse_report_bytes(content, filename="report.xlsx")


def test_workbook_without_property_metrics_is_not_a_report() -> None:
    """A readable spreadsheet still needs a recognizable property-level header."""
    workbook = Workbook()
    workbook.active.append(["Notes", "Value"])
    workbook.active.append(["Status", "Complete"])
    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()

    with pytest.raises(PortfolioManagerParseError, match=r"recognized.*header"):
        parse_report_bytes(buffer.getvalue(), filename="notes.xlsx")


@pytest.mark.parametrize(
    "part_name",
    [
        "xl/vbaProject.bin",
        "xl/externalLinks/externalLink1.xml",
        "xl/embeddings/oleObject1.bin",
        "xl/connections.xml",
    ],
)
def test_ooxml_active_or_external_parts_are_rejected(part_name: str) -> None:
    """A report cannot carry macros, embedded objects, links, or data connections."""
    content = _ooxml_with_part(part_name)

    with pytest.raises(PortfolioManagerParseError, match="external-data OOXML"):
        parse_report_bytes(content, filename="report.xlsx")


def test_duplicate_canonical_headers_are_rejected() -> None:
    """Two aliases for one metric make the report mapping ambiguous."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "Portfolio Manager Property ID",
            "Property ID",
            "Site EUI (kBtu/ft²)",
        ]
    )
    sheet.append(["100", "100", 42])
    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()

    with pytest.raises(PortfolioManagerParseError, match="same metric"):
        parse_report_bytes(buffer.getvalue(), filename="report.xlsx")


def test_oversized_cell_text_is_rejected() -> None:
    """Bounded reports must reject pathological text before result persistence."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Portfolio Manager Property ID", "Property Name", "Site EUI"])
    sheet.append(["100", "x" * (parser._MAX_CELL_TEXT + 1), 42])
    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()

    with pytest.raises(PortfolioManagerParseError, match="text-length"):
        parse_report_bytes(buffer.getvalue(), filename="report.xlsx")


def test_biff_formula_records_are_rejected_before_cached_values_are_read(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Legacy XLS formulas must not be mistaken for trustworthy cached values."""

    class FakeCompoundDocument:
        """Expose one formula record through the small CompDoc surface we use."""

        def __init__(self) -> None:
            """Create a compound document without storage directory entries."""
            self.dirlist = []

        def get_named_stream(self, name: str) -> bytes | None:
            """Return a minimal BIFF stream for the ordinary Workbook name."""
            if name == "Workbook":
                return struct.pack("<HH", 0x0006, 0)
            return None

    monkeypatch.setattr(
        parser,
        "CompDoc",
        lambda *_args, **_kwargs: FakeCompoundDocument(),
    )

    with pytest.raises(PortfolioManagerParseError, match="Formula"):
        parser._preflight_biff(b"compound-document")


def test_biff_macro_storage_is_rejected(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """A legacy workbook containing VBA storage is outside the report contract."""

    class FakeCompoundDocument:
        """Expose a VBA directory entry without requiring an unsafe fixture."""

        def __init__(self) -> None:
            """Create a compound document containing one VBA storage entry."""
            self.dirlist = [SimpleNamespace(name="_VBA_PROJECT")]

        def get_named_stream(self, _name: str) -> bytes | None:
            """No workbook stream should be read after macro detection."""
            return None

    monkeypatch.setattr(
        parser,
        "CompDoc",
        lambda *_args, **_kwargs: FakeCompoundDocument(),
    )

    with pytest.raises(PortfolioManagerParseError, match="Macro-enabled"):
        parser._preflight_biff(b"compound-document")


def test_ooxml_row_limit_is_enforced_without_dimension_metadata(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Streaming iteration must stop even when a worksheet omits max-row metadata."""
    monkeypatch.setattr(parser, "_MAX_ROWS", 1)

    with pytest.raises(PortfolioManagerParseError, match="row limit"):
        parse_report_bytes(_xlsx_bytes(), filename="report.xlsx")


def test_ooxml_column_limit_is_enforced(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """A very wide workbook cannot evade the configured column bound."""
    monkeypatch.setattr(parser, "_MAX_COLUMNS", 2)

    with pytest.raises(PortfolioManagerParseError, match="column limit"):
        parse_report_bytes(_xlsx_bytes(), filename="report.xlsx")


def test_biff_row_limit_is_enforced(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """The genuine legacy reader applies the same bounded-workbook contract."""
    monkeypatch.setattr(parser, "_MAX_ROWS", 1)
    content = (ASSETS / "portfolio-manager-custom-report-single-anonymized.xls").read_bytes()

    with pytest.raises(PortfolioManagerParseError, match="row or column limits"):
        parse_report_bytes(content, filename="report.xls")
