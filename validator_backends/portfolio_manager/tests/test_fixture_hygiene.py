"""Provenance, licensing, and anonymization checks for public fixtures.

Vendored compatibility files are derivative works and may contain workbook
metadata that normal parser assertions never see. These tests keep attribution,
checksums, carrier signatures, hidden content, formulas, links, and known source
identifiers under explicit review.
"""

from __future__ import annotations

import hashlib
import re
import zipfile
from pathlib import Path
from xml.etree import ElementTree

import xlrd
from openpyxl import load_workbook


ASSETS = Path(__file__).with_name("assets")
DATA_FILES = sorted(
    path for path in ASSETS.iterdir() if path.suffix.casefold() in {".xls", ".xlsx", ".xml"}
)
SOURCE_MARKERS = {
    "john skelton",
    "two franklin square",
    "columbia square",
    "5577 13th street",
    "tests@test.com",
    "leslie cook",
    "cook.leslie@epa.gov",
    "school without walls",
    "francis stevens",
    "2425 n street",
    "test city seed",
    "medstar",
    "random person",
    "random@user.net",
    "22178843",
    "22482007",
    "4056933",
}


def _visible_strings(path: Path) -> list[str]:
    """Extract user-visible cells, XML values, and common metadata fields."""
    if path.suffix == ".xlsx":
        workbook = load_workbook(path, read_only=True, data_only=False, keep_links=False)
        try:
            values = [
                str(value)
                for sheet in workbook.worksheets
                for row in sheet.iter_rows(values_only=True)
                for value in row
                if value is not None
            ]
            properties = workbook.properties
            values.extend(
                str(value)
                for value in (
                    properties.creator,
                    properties.lastModifiedBy,
                    properties.title,
                    properties.subject,
                    properties.description,
                )
                if value
            )
            return values
        finally:
            workbook.close()
    if path.suffix == ".xls":
        workbook = xlrd.open_workbook(file_contents=path.read_bytes(), on_demand=True)
        try:
            return [
                str(cell.value)
                for sheet in workbook.sheets()
                for row_index in range(sheet.nrows)
                for cell in sheet.row(row_index)
                if cell.value not in {None, ""}
            ]
        finally:
            workbook.release_resources()
    root = ElementTree.parse(path).getroot()
    return [
        value
        for element in root.iter()
        for value in [element.text, *element.attrib.values()]
        if value
    ]


def test_fixture_checksums_match_reviewed_derivatives() -> None:
    """Unexpected binary fixture changes must force an explicit manifest update."""
    expected = {}
    for line in (ASSETS / "SHA256SUMS").read_text(encoding="utf-8").splitlines():
        digest, name = line.split(maxsplit=1)
        expected[name] = digest

    assert set(expected) == {path.name for path in DATA_FILES}
    for path in DATA_FILES:
        assert hashlib.sha256(path.read_bytes()).hexdigest() == expected[path.name]


def test_fixture_license_and_provenance_are_bundled() -> None:
    """Redistributed derivatives must retain source, revision, and license terms."""
    readme = (ASSETS / "README.md").read_text(encoding="utf-8")
    license_text = (ASSETS / "UPSTREAM-LICENSE.txt").read_text(encoding="utf-8")

    assert "47202f575889f875f78a141ac588b8a7aa7e01eb" in readme
    assert "Anonymization and transformation" in readme
    assert "Alliance for Energy Innovation" in license_text
    assert "Lawrence Berkeley National Laboratory" in license_text


def test_excel_fixtures_use_real_carriers_without_links_or_formulas() -> None:
    """Fixtures must exercise BIFF/OOXML readers without executable workbook content."""
    for path in DATA_FILES:
        if path.suffix == ".xls":
            assert path.read_bytes().startswith(bytes.fromhex("d0cf11e0a1b11ae1"))
        elif path.suffix == ".xlsx":
            assert path.read_bytes().startswith(b"PK\x03\x04")
            with zipfile.ZipFile(path) as archive:
                names = archive.namelist()
                package = b"\n".join(archive.read(name) for name in names)
            assert not any("externalLinks/" in name for name in names)
            assert b"<f" not in package


def test_spreadsheet_fixtures_have_no_hidden_sheets() -> None:
    """Anonymization review must not overlook a hidden worksheet."""
    for path in DATA_FILES:
        if path.suffix != ".xlsx":
            continue
        workbook = load_workbook(path, read_only=True, data_only=False, keep_links=False)
        try:
            assert all(sheet.sheet_state == "visible" for sheet in workbook.worksheets)
        finally:
            workbook.close()


def test_source_identifiers_and_non_fixture_emails_are_absent() -> None:
    """Known upstream identities and contact details cannot survive sanitization."""
    for path in DATA_FILES:
        combined = "\n".join(_visible_strings(path)).casefold()
        assert not any(marker in combined for marker in SOURCE_MARKERS)
        emails = re.findall(r"[\w.+-]+@[\w.-]+\.[a-z]{2,}", combined)
        assert all(email.endswith("@example.invalid") for email in emails)
