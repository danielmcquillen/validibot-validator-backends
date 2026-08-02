"""Bounded carrier readers and canonical Portfolio Manager normalization."""

from __future__ import annotations

import io
import re
import struct
import zipfile
from calendar import monthrange
from datetime import UTC, date, datetime
from decimal import Decimal, InvalidOperation
from html.parser import HTMLParser
from pathlib import PurePosixPath
from typing import TYPE_CHECKING, Any

import xlrd
from defusedxml import ElementTree as DefusedElementTree
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from xlrd.compdoc import CompDoc

from validibot_shared.portfolio_manager import PortfolioManagerPropertyResult


if TYPE_CHECKING:
    from collections.abc import Sequence

type SheetRows = tuple[str, list[list[Any]]]
type HeaderMatch = tuple[int, list[str | None], int]

_MAX_ROWS = 20_000
_MAX_COLUMNS = 500
_MAX_SHEETS = 100
_MAX_XML_ELEMENTS = 200_000
_MAX_XML_DEPTH = 200
_MAX_XML_ATTRIBUTES = 100
_MAX_CELL_TEXT = 4_096
_MAX_OOXML_MEMBERS = 2_000
_MAX_OOXML_MEMBER_BYTES = 100_000_000
_MAX_OOXML_EXPANDED_BYTES = 250_000_000
_MAX_OOXML_COMPRESSION_RATIO = 100
_MAX_MERGED_REGIONS = 10_000
_BIFF_FORMULA_RECORDS = {
    0x0006,
    0x0206,
    0x0406,
    0x0606,
    0x0221,
    0x0236,
    0x04BC,
}
_OOXML_REJECTED_PARTS = (
    "vbaproject.bin",
    "xl/activex/",
    "xl/ctrlprops/",
    "xl/embeddings/",
    "xl/externallinks/",
    "xl/querytables/",
    "xl/connections.xml",
)


class PortfolioManagerParseError(ValueError):
    """A report carrier is safe to read but not a recognized report."""


def _key(value: object) -> str:
    """Normalize report headings and XML names without guessing their meaning."""
    text = _text(value).casefold().replace("²", "2")
    return re.sub(r"[^a-z0-9]+", "", text)


_ALIASES: dict[str, set[str]] = {
    "property_id": {
        "portfoliomanagerpropertyid",
        "portfoliomanagerid",
        "propertyid",
        "propertynumber",
    },
    "property_name": {"propertyname", "name"},
    "parent_property_id": {"parentpropertyid", "parentid"},
    "reporting_period_start": {
        "reportingperiodstartdate",
        "reportingperiodstartingdate",
        "periodstartdate",
        "startdate",
    },
    "reporting_period_end": {
        "reportingperiodenddate",
        "reportingperiodendingdate",
        "periodendingdate",
        "yearending",
        "periodenddate",
        "enddate",
    },
    "property_type": {
        "primarypropertytype",
        "primarypropertytypeselfselected",
        "propertytype",
        "propertytypeselfselected",
        "primaryfunction",
    },
    # EPA's gross floor area excludes parking, so only parking-excluded
    # headings alias onto the canonical field. "Property Floor Area (Buildings
    # and Parking)" and "Property Floor Area (Parking)" are semantically
    # different columns and get their own fields; `_resolve_gross_floor_area`
    # then decides which one backs the canonical value.
    "gross_floor_area_ft2": {
        "propertygfaselfreportedft2",
        "propertygrossfloorareaselfreportedft2",
        "propgrossfloorarea",
        "grossfloorareaft2",
        "grossfloorarea",
        "propertygfa",
    },
    "gross_floor_area_buildings_and_parking_ft2": {
        "propertyfloorareabuildingsandparking",
        "propertyfloorareabuildingsandparkingft2",
    },
    "parking_floor_area_ft2": {
        "propertyfloorareaparking",
        "propertyfloorareaparkingft2",
    },
    "site_eui_kbtu_ft2_yr": {
        "siteeuikbtuft2",
        "siteeuikbtuft2yr",
        "siteintensity",
        "siteeui",
    },
    "weather_normalized_site_eui_kbtu_ft2_yr": {
        "weathernormalizedsiteeuikbtuft2",
        "weathernormalizedsiteeuikbtuft2yr",
        "siteintensitywn",
        "weathernormalizedsiteeui",
        "wneui",
    },
    "source_eui_kbtu_ft2_yr": {
        "sourceeuikbtuft2",
        "sourceeuikbtuft2yr",
        "sourceintensity",
        "sourceeui",
    },
    "national_median_site_eui_kbtu_ft2_yr": {
        "nationalmediansiteeuikbtuft2",
        "nationalmediansiteeuikbtuft2yr",
        "mediansiteintensity",
        "nationalmediansiteeui",
    },
    "site_energy_use_kbtu": {
        "siteenergyusekbtu",
        "sitetotal",
        "siteenergykbtu",
        "siteenergyuse",
    },
    "weather_normalized_site_energy_use_kbtu": {
        "weathernormalizedsiteenergyusekbtu",
        "sitetotalwn",
        "weathernormalizedsiteenergykbtu",
        "weathernormalizedsiteenergyuse",
    },
    "weather_normalized_site_electricity_kwh": {
        "weathernormalizedsiteelectricitykwh",
        "siteelectricitytotalwn",
        "weathernormalizedsiteelectricity",
    },
    "weather_normalized_site_electricity_intensity_kwh_ft2": {
        "weathernormalizedsiteelectricityintensitykwhft2",
        "siteelectricityintensitywn",
        "weathernormalizedsiteelectricityintensity",
    },
    "weather_normalized_site_natural_gas_therms": {
        "weathernormalizedsitenaturalgasusetherms",
        "sitenaturalgasusetotalwn",
        "weathernormalizedsitenaturalgastherms",
        "weathernormalizedsitenaturalgasuse",
    },
    "weather_normalized_site_natural_gas_intensity_therms_ft2": {
        "weathernormalizedsitenaturalgasintensitythermsft2",
        "weathernormalizedsitenaturalgasintensity",
    },
    "onsite_renewable_electricity_generated_kwh": {
        "electricityusegeneratedfromonsiterenewablesystemskwh",
        "electricitygeneratedfromonsiterenewablesystemskwh",
    },
    "onsite_renewable_electricity_exported_kwh": {
        "electricityusegeneratedfromonsiterenewablesystemsandexportedkwh",
        "electricitygeneratedfromonsiterenewablesystemsandexportedkwh",
    },
    "electricity_grid_and_onsite_renewable_kbtu": {
        "electricityusegridpurchaseandgeneratedfromonsiterenewablesystemskbtu",
    },
    "electricity_grid_purchase_kbtu": {
        "electricityusegridpurchasekbtu",
    },
    "onsite_renewable_electricity_used_onsite_kbtu": {
        "electricityusegeneratedfromonsiterenewablesystemsandusedonsitekbtu",
    },
    "natural_gas_use_kbtu": {
        "naturalgasusekbtu",
    },
    "percent_electricity_from_onsite_renewables": {
        "percentoftotalelectricitygeneratedfromonsiterenewablesystems",
        "percentelectricitygeneratedfromonsiterenewablesystems",
    },
    "energy_star_score": {"energystarscore", "score"},
    "heating_degree_days": {"heatingdegreedays", "hdd"},
    "cooling_degree_days": {"coolingdegreedays", "cdd"},
    "weather_station_id": {"weatherstationid"},
    "weather_station_name": {"weatherstationname"},
    "washington_standard_id": {
        "stateofwashingtoncleanbuildingsstandard",
        "washingtoncleanbuildingsstandardid",
    },
}
_ALIAS_TO_FIELD = {alias: field for field, aliases in _ALIASES.items() for alias in aliases}
_DECIMAL_FIELDS = {
    "gross_floor_area_ft2",
    "gross_floor_area_buildings_and_parking_ft2",
    "parking_floor_area_ft2",
    "site_eui_kbtu_ft2_yr",
    "weather_normalized_site_eui_kbtu_ft2_yr",
    "source_eui_kbtu_ft2_yr",
    "national_median_site_eui_kbtu_ft2_yr",
    "site_energy_use_kbtu",
    "weather_normalized_site_energy_use_kbtu",
    "weather_normalized_site_electricity_kwh",
    "weather_normalized_site_electricity_intensity_kwh_ft2",
    "weather_normalized_site_natural_gas_therms",
    "weather_normalized_site_natural_gas_intensity_therms_ft2",
    "onsite_renewable_electricity_generated_kwh",
    "onsite_renewable_electricity_exported_kwh",
    "electricity_grid_and_onsite_renewable_kbtu",
    "electricity_grid_purchase_kbtu",
    "onsite_renewable_electricity_used_onsite_kbtu",
    "natural_gas_use_kbtu",
    "percent_electricity_from_onsite_renewables",
    "energy_star_score",
    "heating_degree_days",
    "cooling_degree_days",
}
_DATE_FIELDS = {"reporting_period_start", "reporting_period_end"}
_MISSING = {"", "n/a", "na", "not available", "not applicable", "--", "none", "null"}
_IDENTITY_FIELDS = {
    "property_id",
    "property_name",
    "parent_property_id",
    "washington_standard_id",
}


def parse_report_bytes(
    content: bytes,
    *,
    filename: str,
) -> list[PortfolioManagerPropertyResult]:
    """Parse one supported report into exactly its canonical property records."""
    suffix = filename.rsplit(".", 1)[-1].casefold() if "." in filename else ""
    if suffix == "xlsx":
        return _records_from_workbook(
            _xlsx_sheets(content),
            filename=filename,
            carrier="xlsx",
        )
    if suffix == "xls":
        prefix = content.lstrip(b"\xef\xbb\xbf \t\r\n").lower()
        if prefix.startswith((b"<?xml", b"<workbook", b"<ss:workbook")):
            rows = _spreadsheetml_rows(content)
            return _records_from_rows(rows, filename=filename, carrier="xls")
        elif prefix.startswith((b"<!doctype html", b"<html", b"<table")):
            rows = _html_table_rows(content)
            return _records_from_rows(rows, filename=filename, carrier="xls")
        elif content.startswith(b"PK\x03\x04"):
            sheets = _xlsx_sheets(content)
        else:
            sheets = _xls_sheets(content)
        return _records_from_workbook(sheets, filename=filename, carrier="xls")
    if suffix == "xml":
        return _xml_records(content, filename=filename)
    raise PortfolioManagerParseError(
        f"Unsupported report extension for {filename!r}; expected .xls, .xlsx, or .xml"
    )


def _xlsx_sheets(content: bytes) -> list[SheetRows]:
    """Read OOXML sheets without trusting optional worksheet dimensions."""
    _preflight_ooxml(content)
    try:
        workbook = load_workbook(
            io.BytesIO(content),
            read_only=True,
            data_only=False,
            keep_links=False,
        )
    except Exception as exc:
        raise PortfolioManagerParseError(f"Could not read OOXML workbook: {exc}") from exc
    try:
        if len(workbook.worksheets) > _MAX_SHEETS:
            raise PortfolioManagerParseError("Workbook exceeds the worksheet limit")
        sheets: list[SheetRows] = []
        for sheet in workbook.worksheets:
            rows: list[list[Any]] = []
            for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                if row_number > _MAX_ROWS:
                    raise PortfolioManagerParseError("Workbook exceeds the row limit")
                values = list(row)
                if len(values) > _MAX_COLUMNS:
                    raise PortfolioManagerParseError("Workbook exceeds the column limit")
                if any(isinstance(value, str) and value.startswith("=") for value in values):
                    raise PortfolioManagerParseError(
                        "Formula cells are not accepted in Portfolio Manager exports"
                    )
                rows.append(values)
            sheets.append((sheet.title, rows))
        return sheets
    finally:
        workbook.close()


def _xls_sheets(content: bytes) -> list[SheetRows]:
    """Read legacy BIFF sheets from bounded in-memory bytes."""
    _preflight_biff(content)
    try:
        workbook = xlrd.open_workbook(file_contents=content, on_demand=True)
    except Exception as exc:
        raise PortfolioManagerParseError(f"Could not read legacy XLS workbook: {exc}") from exc
    sheets: list[SheetRows] = []
    try:
        if workbook.nsheets > _MAX_SHEETS:
            raise PortfolioManagerParseError("Workbook exceeds the worksheet limit")
        for sheet in workbook.sheets():
            if sheet.nrows > _MAX_ROWS or sheet.ncols > _MAX_COLUMNS:
                raise PortfolioManagerParseError("Workbook exceeds row or column limits")
            rows: list[list[Any]] = []
            for row_index in range(sheet.nrows):
                values: list[Any] = []
                for cell in sheet.row(row_index):
                    value = cell.value
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        value = datetime(*xlrd.xldate_as_tuple(value, workbook.datemode))
                    values.append(value)
                rows.append(values)
            sheets.append((sheet.name, rows))
    finally:
        workbook.release_resources()
    return sheets


def _safe_xml_root(content: bytes):
    """Parse XML with entity expansion disabled and explicit shape limits."""
    if re.search(rb"<\?(?!xml(?:\s|\?>))", content, flags=re.IGNORECASE):
        raise PortfolioManagerParseError("XML processing instructions are not accepted")
    try:
        root = DefusedElementTree.fromstring(content)
    except Exception as exc:
        raise PortfolioManagerParseError(f"Could not parse Portfolio Manager XML: {exc}") from exc
    count = 0
    stack = [(root, 1)]
    while stack:
        element, depth = stack.pop()
        count += 1
        if count > _MAX_XML_ELEMENTS:
            raise PortfolioManagerParseError("XML exceeds the element limit")
        if depth > _MAX_XML_DEPTH:
            raise PortfolioManagerParseError("XML exceeds the nesting-depth limit")
        if len(element.attrib) > _MAX_XML_ATTRIBUTES:
            raise PortfolioManagerParseError("XML element exceeds the attribute limit")
        text_values = [element.text, element.tail, *element.attrib.values()]
        if any(len(value or "") > _MAX_CELL_TEXT for value in text_values):
            raise PortfolioManagerParseError("XML text exceeds the value-length limit")
        local_name = _local_name(element.tag).casefold()
        if local_name == "include" and "xinclude" in str(element.tag).casefold():
            raise PortfolioManagerParseError("XML XInclude is not accepted")
        if any(
            _local_name(name).casefold() in {"schemalocation", "nonamespaceschemalocation"}
            for name in element.attrib
        ):
            raise PortfolioManagerParseError("External XML schema locations are not accepted")
        stack.extend((child, depth + 1) for child in list(element))
    return root


def _preflight_ooxml(content: bytes) -> None:
    """Reject active or structurally excessive OOXML before openpyxl reads it."""
    try:
        archive = zipfile.ZipFile(io.BytesIO(content))
    except (OSError, zipfile.BadZipFile) as exc:
        raise PortfolioManagerParseError(f"Could not read OOXML package: {exc}") from exc
    with archive:
        members = [member for member in archive.infolist() if not member.is_dir()]
        if len(members) > _MAX_OOXML_MEMBERS:
            raise PortfolioManagerParseError("OOXML package exceeds the member limit")
        expanded_bytes = 0
        worksheet_count = 0
        seen_names: set[str] = set()
        for member in members:
            name = member.filename
            path = PurePosixPath(name)
            normalized_name = name.casefold()
            if (
                path.is_absolute()
                or ".." in path.parts
                or "\\" in name
                or "\x00" in name
                or normalized_name in seen_names
            ):
                raise PortfolioManagerParseError(
                    "OOXML package contains an unsafe or duplicate part"
                )
            seen_names.add(normalized_name)
            if member.flag_bits & 0x1:
                raise PortfolioManagerParseError("Encrypted OOXML parts are not accepted")
            expanded_bytes += member.file_size
            compressed_size = max(member.compress_size, 1)
            if (
                member.file_size > _MAX_OOXML_MEMBER_BYTES
                or expanded_bytes > _MAX_OOXML_EXPANDED_BYTES
                or member.file_size / compressed_size > _MAX_OOXML_COMPRESSION_RATIO
            ):
                raise PortfolioManagerParseError(
                    "OOXML package exceeds size or compression-ratio limits"
                )
            if any(
                normalized_name == rejected
                or normalized_name.startswith(rejected)
                or (rejected == "vbaproject.bin" and normalized_name.endswith("/vbaproject.bin"))
                for rejected in _OOXML_REJECTED_PARTS
            ):
                raise PortfolioManagerParseError(
                    "Macro, embedded, linked, or external-data OOXML parts are not accepted"
                )
            if normalized_name.startswith("xl/worksheets/") and normalized_name.endswith(".xml"):
                worksheet_count += 1
                worksheet_xml = archive.read(member)
                if len(re.findall(rb"<(?:[A-Za-z0-9_.-]+:)?f(?:\s|>)", worksheet_xml)):
                    raise PortfolioManagerParseError(
                        "Formula cells are not accepted in Portfolio Manager exports"
                    )
                if (
                    len(
                        re.findall(
                            rb"<(?:[A-Za-z0-9_.-]+:)?mergeCell(?:\s|>)",
                            worksheet_xml,
                        )
                    )
                    > _MAX_MERGED_REGIONS
                ):
                    raise PortfolioManagerParseError("Workbook exceeds the merged-region limit")
        if worksheet_count > _MAX_SHEETS:
            raise PortfolioManagerParseError("Workbook exceeds the worksheet limit")


def _preflight_biff(content: bytes) -> None:
    """Reject VBA containers and BIFF formula records before xlrd exposes values."""
    try:
        document = CompDoc(content, logfile=io.StringIO())
    except Exception as exc:
        raise PortfolioManagerParseError(f"Could not read legacy XLS container: {exc}") from exc
    names = {entry.name.casefold() for entry in document.dirlist if entry.name}
    if any("vba" in name or name in {"project", "projectwm", "_vba_project"} for name in names):
        raise PortfolioManagerParseError("Macro-enabled XLS workbooks are not accepted")
    stream = document.get_named_stream("Workbook") or document.get_named_stream("Book")
    if not stream:
        raise PortfolioManagerParseError("Legacy XLS workbook stream is missing")
    position = 0
    while position + 4 <= len(stream):
        opcode, size = struct.unpack_from("<HH", stream, position)
        next_position = position + 4 + size
        if next_position > len(stream):
            raise PortfolioManagerParseError("Legacy XLS record stream is truncated")
        if opcode in _BIFF_FORMULA_RECORDS:
            raise PortfolioManagerParseError(
                "Formula cells are not accepted in Portfolio Manager exports"
            )
        position = next_position


def _spreadsheetml_rows(content: bytes) -> list[list[Any]]:
    """Read the XML Spreadsheet carrier sometimes served with an .xls suffix."""
    root = _safe_xml_root(content)
    rows: list[list[Any]] = []
    for row in root.iter():
        if _local_name(row.tag).casefold() != "row":
            continue
        values: list[Any] = []
        for cell in list(row):
            if _local_name(cell.tag).casefold() != "cell":
                continue
            data = next(
                (child for child in cell.iter() if _local_name(child.tag).casefold() == "data"),
                None,
            )
            values.append(data.text if data is not None else "")
        rows.append(values)
    if not rows:
        raise PortfolioManagerParseError("SpreadsheetML workbook contains no rows")
    return rows


class _BoundedHTMLTableParser(HTMLParser):
    """Extract the first report-like HTML table without executing anything."""

    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.rows: list[list[str]] = []
        self._row: list[str] | None = None
        self._cell_parts: list[str] | None = None

    def handle_starttag(self, tag: str, attrs) -> None:
        """Begin bounded row and cell buffers."""
        del attrs
        normalized = tag.casefold()
        if normalized == "tr":
            self._row = []
        elif normalized in {"td", "th"} and self._row is not None:
            self._cell_parts = []

    def handle_data(self, data: str) -> None:
        """Collect text only while inside a cell."""
        if self._cell_parts is not None:
            self._cell_parts.append(data)

    def handle_endtag(self, tag: str) -> None:
        """Commit complete cells and rows while enforcing parser limits."""
        normalized = tag.casefold()
        if normalized in {"td", "th"} and self._cell_parts is not None:
            value = " ".join(self._cell_parts).strip()
            if len(value) > _MAX_CELL_TEXT:
                raise PortfolioManagerParseError("HTML workbook cell is too large")
            if self._row is not None:
                self._row.append(value)
                if len(self._row) > _MAX_COLUMNS:
                    raise PortfolioManagerParseError("HTML workbook exceeds the column limit")
            self._cell_parts = None
        elif normalized == "tr" and self._row is not None:
            self.rows.append(self._row)
            if len(self.rows) > _MAX_ROWS:
                raise PortfolioManagerParseError("HTML workbook exceeds the row limit")
            self._row = None


def _html_table_rows(content: bytes) -> list[list[Any]]:
    """Read legacy HTML-table workbooks sometimes downloaded with an XLS suffix."""
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            text = content.decode("windows-1252")
        except UnicodeDecodeError as exc:
            raise PortfolioManagerParseError(
                "HTML workbook uses an unsupported character encoding"
            ) from exc
    parser = _BoundedHTMLTableParser()
    try:
        parser.feed(text)
        parser.close()
    except (PortfolioManagerParseError, ValueError) as exc:
        raise PortfolioManagerParseError(f"Could not read HTML workbook: {exc}") from exc
    if not parser.rows:
        raise PortfolioManagerParseError("HTML workbook contains no table rows")
    return parser.rows


def _xml_records(
    content: bytes,
    *,
    filename: str,
) -> list[PortfolioManagerPropertyResult]:
    """Normalize a Portfolio Manager XML response or XML custom report."""
    root = _safe_xml_root(content)
    if _local_name(root.tag).casefold() == "workbook":
        return _records_from_rows(
            _spreadsheetml_rows(content),
            filename=filename,
            carrier="xml",
        )

    report_data_records = []
    for section in root.iter():
        if _key(_local_name(section.tag)) != "informationandmetrics":
            continue
        for element in list(section):
            if _key(_local_name(element.tag)) != "propertymetrics":
                continue
            mapping = _mapping_from_property_metrics(element)
            if mapping.get("property_id"):
                report_data_records.append(
                    _record_from_mapping(
                        mapping,
                        filename=filename,
                        carrier="xml",
                    )
                )
    if report_data_records:
        return report_data_records

    tabular_records = []
    for element in root.iter():
        if _key(_local_name(element.tag)) not in {"row", "record"}:
            continue
        mapping = _mapping_from_xml_row(element)
        if mapping.get("property_id"):
            tabular_records.append(
                _record_from_mapping(
                    mapping,
                    filename=filename,
                    carrier="xml",
                )
            )
    if tabular_records:
        return tabular_records

    candidates = []
    for element in root.iter():
        normalized = _key(_local_name(element.tag))
        if normalized == "propertymetrics":
            mapping = _mapping_from_property_metrics(element)
        elif normalized == "property":
            mapping = _flatten_xml(element)
        else:
            continue
        if mapping.get("property_id"):
            candidates.append((element, mapping))
    if not candidates:
        flattened = _flatten_xml(root)
        if flattened.get("property_id"):
            candidates = [(root, flattened)]
    if not candidates:
        raise PortfolioManagerParseError(
            "XML is not a recognized Portfolio Manager property report"
        )
    return [
        _record_from_mapping(
            mapping,
            filename=filename,
            carrier="xml",
        )
        for _, mapping in candidates
    ]


def _mapping_from_xml_row(element) -> dict[str, Any]:
    """Read generic row/column XML emitted by custom-report download paths."""
    result: dict[str, Any] = {}
    alert_states: dict[str, str] = {}
    custom_ids: dict[str, str] = {}
    for node in element.iter():
        if node is element or list(node):
            continue
        heading = next(
            (
                str(attribute_value)
                for attribute_name, attribute_value in node.attrib.items()
                if _key(_local_name(attribute_name))
                in {"name", "metric", "metricname", "heading", "label"}
            ),
            _local_name(node.tag),
        )
        value = (node.text or "").strip()
        field = _ALIAS_TO_FIELD.get(_key(heading))
        if field:
            result[field] = value
        elif _is_alert_heading(heading):
            alert_states[heading] = _alert_state(value)
        elif _is_identity_heading(heading) and value:
            custom_ids[heading] = value
    result["alert_states"] = alert_states
    result["custom_ids"] = custom_ids
    return result


def _mapping_from_property_metrics(element) -> dict[str, Any]:
    """Read the metric-name/value structure used by report-result XML."""
    result: dict[str, Any] = {
        "property_id": _identity_text(
            next(
                (
                    attribute_value
                    for attribute_name, attribute_value in element.attrib.items()
                    if _key(_local_name(attribute_name))
                    in {"propertyid", "portfoliomanagerpropertyid"}
                ),
                "",
            )
        )
    }
    custom_ids: dict[str, str] = {}
    custom_parts: dict[str, dict[str, str]] = {}
    alert_states: dict[str, str] = {}
    for metric in list(element):
        if _key(_local_name(metric.tag)) != "metric":
            continue
        heading = next(
            (
                str(attribute_value)
                for attribute_name, attribute_value in metric.attrib.items()
                if _key(_local_name(attribute_name))
                in {"name", "metric", "metricname", "heading", "label"}
            ),
            _local_name(metric.tag),
        )
        value_element = next(
            (child for child in list(metric) if _key(_local_name(child.tag)) == "value"),
            None,
        )
        value = (value_element.text or "").strip() if value_element is not None else ""
        custom_match = re.fullmatch(
            r"custompropertyid(\d+)(name|id|number|value)",
            _key(heading),
        )
        field = _ALIAS_TO_FIELD.get(_key(heading))
        if field and (value or field not in result):
            result[field] = value
        elif _is_alert_heading(heading):
            alert_states[heading] = _alert_state(value)
        elif _is_identity_heading(heading) and value and custom_match is None:
            custom_ids[heading] = value

        if custom_match and value:
            slot, part = custom_match.groups()
            normalized_part = "name" if part == "name" else "value"
            custom_parts.setdefault(slot, {})[normalized_part] = value

    for pair in custom_parts.values():
        name = _text(pair.get("name"))
        value = _identity_text(pair.get("value"))
        if name and value:
            custom_ids[name] = value
            if _key(name) == "stateofwashingtoncleanbuildingsstandard":
                result["washington_standard_id"] = value

    if not _text(result.get("reporting_period_end")):
        try:
            year = int(element.attrib.get("year", ""))
            month = int(element.attrib.get("month", ""))
            result["reporting_period_end"] = date(
                year,
                month,
                monthrange(year, month)[1],
            )
        except (TypeError, ValueError):
            pass
    result["custom_ids"] = custom_ids
    result["alert_states"] = alert_states
    return result


def _flatten_xml(element) -> dict[str, Any]:
    """Flatten recognized metric leaves and named IDs below one property node."""
    result: dict[str, Any] = {}
    custom_ids: dict[str, str] = {}
    for node in element.iter():
        children = list(node)
        local = _local_name(node.tag)
        normalized = _key(local)
        value = (node.text or "").strip()
        if children and normalized not in {"standardid", "customid"}:
            continue
        field = _ALIAS_TO_FIELD.get(normalized)
        if field and value:
            result[field] = value
        if normalized in {"standardid", "customid"}:
            name = next(
                (
                    str(attribute_value)
                    for attribute_name, attribute_value in node.attrib.items()
                    if _key(_local_name(attribute_name)) in {"name", "type"}
                ),
                "",
            )
            id_value = value or next(
                (
                    (child.text or "").strip()
                    for child in children
                    if _key(_local_name(child.tag)) in {"id", "value"}
                ),
                "",
            )
            if name and id_value:
                custom_ids[name] = id_value
                if _key(name) == "stateofwashingtoncleanbuildingsstandard":
                    result["washington_standard_id"] = id_value
    result["custom_ids"] = custom_ids
    return result


def _records_from_workbook(
    sheets: Sequence[SheetRows],
    *,
    filename: str,
    carrier: str,
) -> list[PortfolioManagerPropertyResult]:
    """Select one property-level report sheet without flattening other tables."""
    candidates: list[tuple[tuple[int, int, int, int], int, HeaderMatch]] = []
    for sheet_index, (sheet_name, rows) in enumerate(sheets):
        match = _find_report_header(rows)
        if match is None:
            continue
        _, mapped_headers, known_count = match
        fields = {field for field in mapped_headers if field}
        evidence_count = len(fields - _IDENTITY_FIELDS)
        if evidence_count == 0 and "washington_standard_id" not in fields:
            continue
        normalized_name = _key(sheet_name)
        preferred = {
            "informationandmetrics": 2,
            "property": 1,
        }.get(normalized_name, 0)
        candidates.append(
            (
                (preferred, evidence_count, known_count, -sheet_index),
                sheet_index,
                match,
            )
        )
    if not candidates:
        raise PortfolioManagerParseError(
            "Workbook does not contain a recognized Portfolio Manager report header"
        )

    _, primary_sheet_index, primary_match = max(candidates, key=lambda item: item[0])
    records = _records_from_rows(
        sheets[primary_sheet_index][1],
        filename=filename,
        carrier=carrier,
        header_match=primary_match,
    )
    property_ids_sheet = next(
        (rows for sheet_name, rows in sheets if _key(sheet_name) == "propertyids"),
        None,
    )
    if property_ids_sheet:
        records = _merge_property_ids(records, property_ids_sheet)
    return records


def _find_report_header(rows: Sequence[Sequence[Any]]) -> HeaderMatch | None:
    """Find a property header inside one worksheet's leading report metadata."""
    for index, row in enumerate(rows[:100]):
        candidates = [_ALIAS_TO_FIELD.get(_key(value)) for value in row]
        known = sum(candidate is not None for candidate in candidates)
        if "property_id" in candidates and known >= 2:
            known_fields = [candidate for candidate in candidates if candidate]
            if len(known_fields) != len(set(known_fields)):
                raise PortfolioManagerParseError(
                    "Workbook report header maps more than one column to the same metric"
                )
            return index, candidates, known
    return None


def _merge_property_ids(
    records: list[PortfolioManagerPropertyResult],
    rows: Sequence[Sequence[Any]],
) -> list[PortfolioManagerPropertyResult]:
    """Join the full-export Property IDs sheet onto primary property records."""
    match = _find_report_header(rows)
    if match is None:
        return records
    header_index, mapped_headers, _ = match
    headers = rows[header_index]
    identities: dict[str, tuple[dict[str, str], str]] = {}
    for row in rows[header_index + 1 :]:
        property_id = ""
        for column, field in enumerate(mapped_headers):
            if field == "property_id" and column < len(row):
                property_id = _identity_text(row[column])
                break
        if not property_id:
            continue
        custom_ids: dict[str, str] = {}
        washington_standard_id = ""
        column = 0
        while column < len(headers):
            heading = _text(headers[column])
            normalized = _key(heading)
            raw = row[column] if column < len(row) else None
            value = _identity_text(raw)
            if normalized in {"customidname", "custompropertyidname"}:
                number_column = column + 1
                number_heading = (
                    _key(headers[number_column]) if number_column < len(headers) else ""
                )
                number_value = (
                    _identity_text(row[number_column]) if number_column < len(row) else ""
                )
                if number_heading in {
                    "customidnumber",
                    "custompropertyidnumber",
                    "customidid",
                    "custompropertyidid",
                }:
                    if value and number_value:
                        custom_ids[value] = number_value
                    column += 2
                    continue
            if (
                column >= 2
                and heading
                and value
                and normalized not in {"customidnumber", "custompropertyidnumber"}
            ):
                custom_ids[heading] = value
                if normalized == "stateofwashingtoncleanbuildingsstandard":
                    washington_standard_id = value
            column += 1
        identities[property_id] = (custom_ids, washington_standard_id)

    merged: list[PortfolioManagerPropertyResult] = []
    for record in records:
        custom_ids, washington_standard_id = identities.get(record.property_id, ({}, ""))
        if not custom_ids and not washington_standard_id:
            merged.append(record)
            continue
        merged.append(
            record.model_copy(
                update={
                    "custom_ids": {**record.custom_ids, **custom_ids},
                    "washington_standard_id": (
                        washington_standard_id or record.washington_standard_id
                    ),
                }
            )
        )
    return merged


def _records_from_rows(
    rows: Sequence[Sequence[Any]],
    *,
    filename: str,
    carrier: str,
    header_match: HeaderMatch | None = None,
) -> list[PortfolioManagerPropertyResult]:
    """Discover a metric header and normalize every nonblank data row below it."""
    match = header_match or _find_report_header(rows)
    if match is None:
        raise PortfolioManagerParseError(
            "Workbook does not contain a recognized Portfolio Manager report header"
        )
    header_index, mapped_headers, _ = match

    records: list[PortfolioManagerPropertyResult] = []
    for row in rows[header_index + 1 :]:
        if not any(_text(value) for value in row):
            continue
        mapping: dict[str, Any] = {}
        alert_states: dict[str, str] = {}
        for column, field in enumerate(mapped_headers):
            if column >= len(row):
                continue
            raw = row[column]
            header_text = str(rows[header_index][column] or "")
            if field:
                mapping[field] = raw
            elif _is_alert_heading(header_text):
                alert_states[header_text.strip()] = _alert_state(raw)
            elif _is_identity_heading(header_text) and _identity_text(raw):
                mapping.setdefault("custom_ids", {})[header_text.strip()] = _identity_text(raw)
        mapping["alert_states"] = alert_states
        if not _identity_text(mapping.get("property_id")):
            continue
        if _key(mapping["property_id"]) in _ALIASES["property_id"]:
            continue
        records.append(_record_from_mapping(mapping, filename=filename, carrier=carrier))
    if not records:
        raise PortfolioManagerParseError("Portfolio Manager report contains no property row")
    return records


def _record_from_mapping(
    mapping: dict[str, Any],
    *,
    filename: str,
    carrier: str,
) -> PortfolioManagerPropertyResult:
    """Coerce one carrier mapping into the canonical property model."""
    values: dict[str, Any] = {
        "member_name": filename,
        "carrier": carrier,
        "property_id": _identity_text(mapping.get("property_id")),
        "property_name": _text(mapping.get("property_name")),
        "parent_property_id": _identity_text(mapping.get("parent_property_id")),
        "property_type": _text(mapping.get("property_type")),
        "weather_station_id": _text(mapping.get("weather_station_id")),
        "weather_station_name": _text(mapping.get("weather_station_name")),
        "washington_standard_id": _identity_text(mapping.get("washington_standard_id")),
        "custom_ids": {
            _text(name): _identity_text(value)
            for name, value in (mapping.get("custom_ids") or {}).items()
            if _text(name) and _identity_text(value)
        },
        "alert_states": mapping.get("alert_states") or {},
    }
    metric_states: dict[str, str] = {}
    for field in _DECIMAL_FIELDS:
        raw = mapping.get(field)
        parsed_decimal = _decimal(raw)
        values[field] = parsed_decimal
        if field not in mapping:
            metric_states[field] = "absent"
        elif parsed_decimal is not None:
            metric_states[field] = "value"
        elif _text(raw).casefold() in _MISSING:
            metric_states[field] = "not_available"
        else:
            metric_states[field] = "invalid"
    _resolve_gross_floor_area(values, metric_states)
    values["metric_states"] = metric_states
    for field in _DATE_FIELDS:
        raw = mapping.get(field)
        parsed_date = _date(raw)
        values[field] = parsed_date
        if field not in mapping:
            metric_states[field] = "absent"
        elif parsed_date is not None:
            metric_states[field] = "value"
        elif _text(raw).casefold() in _MISSING:
            metric_states[field] = "not_available"
        else:
            metric_states[field] = "invalid"
    if not values["property_id"]:
        raise PortfolioManagerParseError("A property row is missing Portfolio Manager Property ID")
    return PortfolioManagerPropertyResult.model_validate(values)


def _resolve_gross_floor_area(
    values: dict[str, Any],
    metric_states: dict[str, str],
) -> None:
    """Back the canonical GFA with a parking-excluded value and record its basis.

    Portfolio Manager publishes three related floor-area columns, and only the
    self-reported one matches EPA's gross floor area definition (and therefore
    Washington's Form B basis). Precedence is explicit so the resolved value
    never depends on the order headings or XML metrics happen to appear in:

    1. a directly reported parking-excluded column wins outright;
    2. otherwise the parking-inclusive column minus the parking column is the
       parking-excluded area, by EPA's own definition of the two columns;
    3. otherwise the parking-inclusive column is used unchanged, because many
       report templates carry no other floor area at all — but the basis says
       so, rather than letting a parking-inclusive number pass as GFA.
    """
    self_reported = values["gross_floor_area_ft2"]
    inclusive = values["gross_floor_area_buildings_and_parking_ft2"]
    parking = values["parking_floor_area_ft2"]

    if self_reported is not None:
        basis = "self_reported"
    elif inclusive is None:
        basis = "absent"
    elif parking is not None and inclusive >= parking:
        # Parking never exceeds buildings-and-parking in a coherent report; a
        # contradiction means the two columns cannot be reconciled, so fall
        # through to the unchanged inclusive value rather than invent one.
        values["gross_floor_area_ft2"] = inclusive - parking
        basis = "buildings_and_parking_less_parking"
    else:
        values["gross_floor_area_ft2"] = inclusive
        basis = "buildings_and_parking"

    values["gross_floor_area_basis"] = basis
    if values["gross_floor_area_ft2"] is not None:
        metric_states["gross_floor_area_ft2"] = "value"
    elif metric_states["gross_floor_area_ft2"] == "absent":
        # No parking-excluded column was reported, so the canonical metric is
        # only as knowable as the parking-inclusive column it would derive from.
        metric_states["gross_floor_area_ft2"] = metric_states[
            "gross_floor_area_buildings_and_parking_ft2"
        ]


def _decimal(value: Any) -> Decimal | None:
    """Parse Portfolio Manager numeric display values without binary floats."""
    text = _text(value)
    if text.casefold() in _MISSING:
        return None
    cleaned = (
        text.replace(",", "")
        .replace("$", "")
        .replace("%", "")
        .replace("\N{MINUS SIGN}", "-")
        .strip()
    )
    match = re.match(r"^[+-]?(?:\d+(?:\.\d*)?|\.\d+)", cleaned)
    if not match:
        return None
    try:
        return Decimal(match.group(0))
    except InvalidOperation:
        return None


def _date(value: Any) -> date | None:
    """Parse native spreadsheet dates and common Portfolio Manager displays."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and value > 0:
        try:
            converted = from_excel(value)
            return converted.date() if isinstance(converted, datetime) else converted
        except (TypeError, ValueError, OverflowError):
            pass
    text = _text(value)
    if text.casefold() in _MISSING:
        return None
    for pattern in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%b %d, %Y", "%Y"):
        try:
            parsed = datetime.strptime(text, pattern).replace(tzinfo=UTC)
            if pattern == "%Y":
                return date(parsed.year, 12, 31)
            return parsed.date()
        except ValueError:
            continue
    return None


def _alert_state(value: Any) -> str:
    """Normalize report alert cells without treating absence as a clean result."""
    text = _text(value).casefold()
    if not text:
        return "not_verifiable"
    if text in {"n/a", "na", "not available", "not applicable", "--"}:
        return "not_verifiable"
    if text in {"no", "none", "false", "0", "ok", "pass", "passed"}:
        return "clean"
    return "alert"


def _is_identity_heading(value: object) -> bool:
    """Recognize identity columns without treating every unknown metric as an ID."""
    normalized = _key(value)
    return (
        normalized.startswith("customid")
        or normalized.startswith("custompropertyid")
        or normalized.startswith("standardid")
    )


def _is_alert_heading(value: object) -> bool:
    """Recognize EPA Alert Metrics and the estimated-energy evidence field."""
    normalized = _key(value)
    return "alert" in normalized or (
        "estimated" in normalized
        and any(token in normalized for token in ("energy", "electricity", "naturalgas"))
    )


def _text(value: Any) -> str:
    """Return trimmed display text while preserving opaque identifiers."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        text = str(int(value))
    else:
        text = str(value).strip()
    if len(text) > _MAX_CELL_TEXT:
        raise PortfolioManagerParseError("Report value exceeds the text-length limit")
    return text


def _identity_text(value: Any) -> str:
    """Normalize opaque IDs while removing Portfolio Manager absence labels."""
    text = _text(value)
    normalized = text.casefold()
    if normalized in _MISSING or normalized.startswith(("not applicable:", "not available:")):
        return ""
    return text


def _local_name(tag: str) -> str:
    """Strip Clark-notation or prefixed XML namespaces."""
    return tag.rsplit("}", 1)[-1].rsplit(":", 1)[-1]
